from django.shortcuts import render, redirect

from noticias.models import Noticia
from .forms import EncuestaForm
from django.views.generic import TemplateView

# Create your views here.
from .models import Encuestavacunacion
from openpyxl import Workbook
from django.http.response import HttpResponse


def encuestas(request):
    if request.method == 'POST':
        formaencuesta = EncuestaForm(request.POST)
        if formaencuesta.is_valid():
            formaencuesta.save()
            return redirect('http://127.0.0.1:8000/')

    else:
        formaencuesta = EncuestaForm()

    return render(request, 'encuesta.html', {'formaencuesta': formaencuesta})


def lista(request):
    lista = Encuestavacunacion.objects.order_by('id')
    return render(request, 'lista.html', {'lista': lista})


class reporteencuesta(TemplateView):
    def get(self, request, *args, **kwargs):
        lista = Encuestavacunacion.objects.order_by('id')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE ENCUESTAS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'id'
        ws['C3'] = 'nombre'
        ws['D3'] = 'apellido'
        ws['E3'] = 'edad'
        ws['F3'] = 'entidad'
        ws['G3'] = 'vacunado'
        ws['H3'] = 'fecha'
        ws['I3'] = 'genero'
        cont = 4
        for encuesta in lista:
            ws.cell(row=cont, column=2).value = encuesta.id
            ws.cell(row=cont, column=3).value = encuesta.nombre
            ws.cell(row=cont, column=4).value = encuesta.apellido
            ws.cell(row=cont, column=5).value = encuesta.edad
            ws.cell(row=cont, column=6).value = encuesta.entidad.estado
            ws.cell(row=cont, column=7).value = encuesta.vacunado.vacuna
            ws.cell(row=cont, column=8).value = encuesta.fecha
            ws.cell(row=cont, column=9).value = encuesta.genero.sexo
            cont += 1

        nombre_Archivo = "listaEncuestasCovid.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_Archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
